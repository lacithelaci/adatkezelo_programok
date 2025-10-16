from typing import Type, cast
import openpyxl
from openpyxl.workbook import Workbook

from data.basic.model_dataclasses import Person, Movie, Review


def write_people(people: list[Person], wb: Workbook,
                 sheet_name: str = "people",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["person_id", "name", "age", "city", "male"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(people)):
        sheet.cell(row=row + offset, column=1, value=people[row].person_id)
        sheet.cell(row=row + offset, column=2, value=people[row].name)
        sheet.cell(row=row + offset, column=3, value=people[row].age)
        sheet.cell(row=row + offset, column=4, value=people[row].city)
        sheet.cell(row=row + offset, column=5, value=people[row].male)


def read_people(wb: Workbook,
                sheet_name: str = "people",
                heading: bool = True) -> list[Person]:
    sheet = wb[sheet_name]

    people = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        people.append(
            Person(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value,
                sheet.cell(row=row, column=5).value
            )
        )
        row += 1
    return people


def write_movie(movie: list[Movie], wb: Workbook,
                sheet_name: str = "movie",
                heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["movie_id", "title", "year", "genre"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(movie)):
        sheet.cell(row=row + offset, column=1, value=movie[row].movie_id)
        sheet.cell(row=row + offset, column=2, value=movie[row].title)
        sheet.cell(row=row + offset, column=3, value=movie[row].year)
        sheet.cell(row=row + offset, column=4, value=movie[row].genre)


def read_movie(wb: Workbook,
               sheet_name: str = "movie",
               heading: bool = True) -> list[Movie]:
    sheet = wb[sheet_name]

    movie = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        movie.append(
            Movie(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value
            )
        )
        row += 1
    return movie


def write_review(review: list[Review], wb: Workbook,
                 sheet_name: str = "review",
                 heading: bool = True) -> None:
    sheet = wb.create_sheet(sheet_name)

    if heading:
        column_names = ["review_id", "person_id", "movie_id", "rating", "text", "date"]
        for col in range(len(column_names)):
            sheet.cell(row=1, column=col + 1, value=column_names[col])

    offset = 2 if heading else 1
    for row in range(len(review)):
        sheet.cell(row=row + offset, column=1, value=review[row].review_id)
        sheet.cell(row=row + offset, column=2, value=review[row].person_id)
        sheet.cell(row=row + offset, column=3, value=review[row].movie_id)
        sheet.cell(row=row + offset, column=4, value=review[row].rating)
        sheet.cell(row=row + offset, column=5, value=review[row].text)
        sheet.cell(row=row + offset, column=6, value=review[row].date)


def read_review(wb: Workbook,
                sheet_name: str = "review",
                heading: bool = True) -> list[Review]:
    sheet = wb[sheet_name]

    review = []
    row = 2 if heading else 1
    while True:
        cell = sheet.cell(row=row, column=1)
        if cell.value is None:
            break

        review.append(
            Review(
                sheet.cell(row=row, column=1).value,
                sheet.cell(row=row, column=2).value,
                sheet.cell(row=row, column=3).value,
                sheet.cell(row=row, column=4).value,
                sheet.cell(row=row, column=5).value,
                sheet.cell(row=row, column=6).value
            )
        )
        row += 1
    return review

def read(entity_type: Type[object], workbook: openpyxl.Workbook,
         sheet_name: str = None, heading: bool = True) -> list[object]:

    if sheet_name is None:
        if entity_type == Person:
            sheet_name = "people"
        elif entity_type == Movie:
            sheet_name = "movie"
        elif entity_type == Review:
            sheet_name = "review"
        else:
            raise RuntimeError("Unknown type of entity")

    if entity_type == Person:
        return read_people(workbook, sheet_name=sheet_name, heading=heading)
    elif entity_type == Movie:
        return read_movie(workbook, sheet_name=sheet_name, heading=heading)
    elif entity_type == Review:
        return read_review(workbook, sheet_name=sheet_name, heading=heading)


def write(entities: list[object], workbook: openpyxl.Workbook,
          sheet_name: str = None, heading: bool = True) -> None:
    if sheet_name is None:
        if isinstance(entities[0], Person):
            sheet_name = "people"
        elif isinstance(entities[0], Movie):
            sheet_name = "movie"
        elif isinstance(entities[0], Review):
            sheet_name = "review"
        else:
            raise RuntimeError("Unknown type of entity")

    if isinstance(entities[0], Person):
        return write_people([cast(Person, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Movie):
        return write_movie([cast(Movie, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
    elif isinstance(entities[0], Review):
        return write_review([cast(Review, e) for e in entities], workbook, sheet_name=sheet_name, heading=heading)
